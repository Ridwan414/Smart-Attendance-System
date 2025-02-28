import cv2
import os
import time
import numpy as np
import face_recognition
import joblib
import pandas as pd
from flask import Flask, request, render_template, redirect, url_for
from datetime import datetime
from mtcnn import MTCNN
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Paths
KNOWN_FACES_DIR = "static/faces"
ATTENDANCE_CSV = f"Attendance/Attendance-{datetime.today().strftime('%m_%d_%y')}.csv"

# Create directories if they don’t exist
os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
os.makedirs("Attendance", exist_ok=True)

# Initialize MTCNN
detector = MTCNN()

# Face recognition threshold
FACE_MATCH_THRESHOLD = 0.45  # Stricter threshold for higher accuracy

# Load saved encodings
if os.path.exists("face_encodings.pkl"):
    known_face_encodings, known_face_names = joblib.load("face_encodings.pkl")
else:
    known_face_encodings, known_face_names = [], []


# Function to load known faces
def load_known_faces():
    global known_face_encodings, known_face_names
    known_face_encodings = []
    known_face_names = []

    # If no faces exist, clear encodings
    if not os.listdir(KNOWN_FACES_DIR):
        joblib.dump(([], []), "face_encodings.pkl")
        print("⚠️ No registered faces found. Resetting encodings.")
        return

    # Load new faces if available
    for person in os.listdir(KNOWN_FACES_DIR):
        person_dir = os.path.join(KNOWN_FACES_DIR, person)
        if os.path.isdir(person_dir):
            for filename in os.listdir(person_dir):
                img_path = os.path.join(person_dir, filename)
                image = face_recognition.load_image_file(img_path)
                encodings = face_recognition.face_encodings(image)

                if encodings:
                    known_face_encodings.append(encodings[0])
                    known_face_names.append(person)

    joblib.dump((known_face_encodings, known_face_names), "face_encodings.pkl")
    print(f"✅ Loaded {len(known_face_encodings)} registered faces.")



# Load faces if not already loaded
if not known_face_encodings:
    load_known_faces()


# Function to detect and recognize faces using MTCNN
def recognize_faces(frame):
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    detections = detector.detect_faces(rgb_frame)

    recognized_names = []
    for detection in detections:
        x, y, width, height = detection['box']
        face_location = (y, x + width, y + height, x)

        face_encoding = face_recognition.face_encodings(rgb_frame, [face_location])
        if not face_encoding:
            continue

        encoding = face_encoding[0]
        matches = face_recognition.compare_faces(known_face_encodings, encoding, tolerance=FACE_MATCH_THRESHOLD)
        name = "Unknown"

        if True in matches:
            match_index = np.argmin(face_recognition.face_distance(known_face_encodings, encoding))
            name = known_face_names[match_index]

        recognized_names.append((name, (x, y, x + width, y + height)))

    return recognized_names

# Global flag to track if a scan session has been started
session_started = False  # Track whether a scan session has started
current_wb = None
current_sheet = None
current_file = None

# Function to mark attendance
def mark_attendance(name):
    global session_started, current_wb, current_sheet, current_file
    if name == "Unknown":
        return

    try:
        # Ensure the Attendance folder exists
        os.makedirs("Attendance", exist_ok=True)

        # Generate CSV attendance file path with today's date
        attendance_file = f"Attendance/Attendance-{datetime.today().strftime('%m_%d_%y')}.csv"

        # Generate Excel attendance file path with today's date
        excel_file = f"Attendance/Attendance-{datetime.today().strftime('%d_%m_%Y')}.xlsx"

        # If session is not started yet, initialize it
        if not session_started:
            start_process()  # Initialize the Excel process
            session_started = True

        # Check if CSV file exists and load existing data
        if os.path.exists(attendance_file):
            df = pd.read_csv(attendance_file)
        else:
            df = pd.DataFrame(columns=["User ID", "Name", "Time"])

        current_time = datetime.now().strftime("%H:%M:%S")

        # Extract user ID from stored name format
        user_id = name.split("_")[-1]

        # Check if already marked in the CSV
        if user_id not in df["User ID"].values:
            new_entry = pd.DataFrame({"User ID": [user_id], "Name": [name], "Time": [current_time]})
            df = pd.concat([df, new_entry], ignore_index=True)

            # Save and flush data immediately to CSV
            df.to_csv(attendance_file, index=False)
            print(f"✅ Attendance marked for {name} (User ID: {user_id}) at {current_time}")

            # Add to the Excel sheet
            add_to_excel(name, user_id, current_time)

        else:
            print(f"⚠️ {name} (User ID: {user_id}) is already marked present.")

    except Exception as e:
        print(f"❌ Error marking attendance: {e}")


# Excel sheet related code
def get_today_filename():
    today = datetime.now().strftime('%d_%m_%Y')
    return f"Attendance-{today}.xlsx"


def start_process():
    global current_wb, current_sheet, current_file
    try:
        current_file = get_today_filename()
        if os.path.exists(f'Attendance/{current_file}'):
            current_wb = load_workbook(f'Attendance/{current_file}')
            sheet_count = len(current_wb.sheetnames)
        else:
            current_wb = Workbook()
            default_sheet = current_wb.active
            current_wb.remove(default_sheet)
            sheet_count = 0

        period_name = get_period_name(sheet_count + 1)

        current_sheet = current_wb.create_sheet(period_name)
        current_sheet['A1'] = "Name"
        current_sheet['B1'] = "ID"
        current_sheet['C1'] = "Time"

        current_wb.save(f'Attendance/{current_file}')
        print(f"✅ New scan session started, new sheet created: {period_name}")
    except Exception as e:
        print(f"Error starting process: {e}")

def get_period_name(index):
    if index % 10 == 1 and index % 100 != 11:
        suffix = "st"
    elif index % 10 == 2 and index % 100 != 12:
        suffix = "nd"
    elif index % 10 == 3 and index % 100 != 13:
        suffix = "rd"
    else:
        suffix = "th"
    return f"{index}{suffix} Period"


# Function to add attendance to the Excel sheet
def add_to_excel(name, user_id, current_time):
    try:
        if current_wb is None:
            start_process()

# Check if the user has already marked attendance
        existing_user_ids = [row[1] for row in current_sheet.iter_rows(min_row=2, values_only=True)]  # Extract all User IDs
        if user_id in existing_user_ids:
            print(f"⚠️ {name} (User ID: {user_id}) has already marked attendance. Skipping...")
            return  # If the user already exists, don't add them again

# If the user hasn't marked attendance yet, add the new row
        new_row = [name, user_id, current_time]
        current_sheet.append(new_row)

# Save the changes to the Excel file
        current_wb.save(f'Attendance/{current_file}')
        print(f"✅ Attendance added to Excel for {name} (User ID: {user_id}) at {current_time}")
    except Exception as e:
        print(f"❌ Error adding to Excel: {e}")


# Function to close the session when Esc is pressed
def close_session():
    global session_started, current_wb
    if session_started:
        try:
            current_wb.save(f'Attendance/{current_file}')
            print(f"✅ Session closed, data saved in {current_file}")
            session_started = False
        except Exception as e:
            print(f"❌ Error closing session: {e}")

#Define the path for the attendance folder
ATTENDANCE_DIR = 'Attendance'
# This function retrieves the latest Excel file based on the current date
def get_latest_excel_file():
    today_filename = f"Attendance-{datetime.today().strftime('%d_%m_%Y')}.xlsx"
    file_path = os.path.join(ATTENDANCE_DIR, today_filename)
    if os.path.exists(file_path):
        return file_path
    return None

# Function to get the latest sheet's data from the Excel file
def get_latest_sheet_data(excel_file):
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        # Get the latest sheet, i.e., the last sheet in the workbook
        latest_sheet = wb[wb.sheetnames[-1]]
        
        # Read the sheet's content into a pandas DataFrame
        data = []
        for row in latest_sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip header
            data.append({"Name": row[0], "User ID": row[1], "Time": row[2]})
        
        return data
    except Exception as e:
        print(f"❌ Error reading the latest sheet: {e}")
        return []

# Flask Routes

@app.route("/")
def home():
    # Get the latest Excel file
    latest_excel_file = get_latest_excel_file()
    
    # Get the data from the latest sheet
    if latest_excel_file:
        attendance_data = get_latest_sheet_data(latest_excel_file)
    else:
        attendance_data = []

    # Count total registered users (assuming you have a separate folder for known faces)
    total_users = len(os.listdir(KNOWN_FACES_DIR)) if os.path.exists(KNOWN_FACES_DIR) else 0

    return render_template(
        "home.html",
        total_users=total_users,  # Send total user count
        attendance=attendance_data  # Send attendance data from the latest sheet
    )



@app.route("/add", methods=["POST"])
def register_user():
    if "username" not in request.form or "user_id" not in request.form:
        return render_template("home.html", message="Error: Username and User ID required!")

    username = request.form["username"].strip()
    user_id = request.form["user_id"].strip()

    if not username or not user_id:
        return render_template("home.html", message="Username and User ID cannot be empty!")

    user_dir = os.path.join(KNOWN_FACES_DIR, f"{username}_{user_id}")
    os.makedirs(user_dir, exist_ok=True)

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        return render_template("home.html", message="Error: Could not access webcam!")

    print(f"📸 Please position your face within the frame for {username} (User ID: {user_id})...")

    face_detected_frames = 0  # Counter for correctly positioned frames
    required_frames = 3  # User must hold still for 10 frames

    while True:
        ret, frame = cap.read()
        if not ret:
            continue

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        detections = detector.detect_faces(rgb_frame)

        h, w, _ = frame.shape
        guide_x1, guide_y1, guide_x2, guide_y2 = int(w * 0.3), int(h * 0.3), int(w * 0.7), int(h * 0.7)

        # Draw a guiding box for face alignment
        cv2.rectangle(frame, (guide_x1, guide_y1), (guide_x2, guide_y2), (255, 255, 0), 2)
        cv2.putText(frame, "Align your face within the blue box", (50, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)

        face_properly_aligned = False

        for detection in detections:
            x, y, width, height = detection['box']
            x2, y2 = x + width, y + height

            # Draw a bounding box around the detected face
            cv2.rectangle(frame, (x, y), (x2, y2), (0, 255, 0), 2)

            # Face is properly positioned if within the guiding box
            if guide_x1 < x and guide_x2 > x2 and guide_y1 < y and guide_y2 > y2 and 120 < width < 300:
                face_properly_aligned = True
                face_detected_frames += 1
            else:
                face_detected_frames = 0  # Reset counter if face moves out of position

            # Display progress bar
            progress_width = int((face_detected_frames / required_frames) * 300)
            cv2.rectangle(frame, (50, h - 50), (50 + progress_width, h - 30), (0, 255, 0), -1)
            cv2.rectangle(frame, (50, h - 50), (350, h - 30), (255, 255, 255), 2)
            cv2.putText(frame, f"Hold still: {face_detected_frames}/{required_frames}", (50, h - 60),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        cv2.imshow("Registering User - Adjust Your Position", frame)

        # If face is properly aligned for the required frames, capture images
        if face_detected_frames >= required_frames:
            print("✅ Face positioned correctly. Capturing images...")
            break

        # Press ESC to cancel registration
        if cv2.waitKey(1) & 0xFF == 27:
            print("🚪 User cancelled registration.")
            cap.release()
            cv2.destroyAllWindows()
            return render_template("home.html", message="Registration cancelled.")

    # Automatically capture 5 images with a delay
    for i in range(5):
        ret, frame = cap.read()
        if not ret:
            print(f"❌ Failed to capture frame {i + 1}")
            continue

        # Display capture count
        cv2.putText(frame, f"Capturing Image {i + 1}/5", (50, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
        cv2.imshow("Registering User", frame)
        cv2.imwrite(os.path.join(user_dir, f"{username}_{user_id}_{i}.jpg"), frame)

        time.sleep(6)  # Delay between captures
        if cv2.waitKey(1) & 0xFF == 27:
            print("🚪 User cancelled registration.")
            cap.release()
            cv2.destroyAllWindows()
            return render_template("home.html", message="Registration cancelled.")

    cap.release()
    cv2.destroyAllWindows()

    print(f"✅ Successfully registered {username} (User ID: {user_id})!")

    # Load newly added faces
    load_known_faces()

    return render_template("home.html", message=f"✅ {username} (User ID: {user_id}) successfully registered with 5 images!")





@app.route("/attendance/faces")
def start_attendance():
    global known_face_encodings, known_face_names

    # Reload faces in case they were cleared
    if not os.listdir(KNOWN_FACES_DIR):
        known_face_encodings, known_face_names = [], []
        print("⚠️ No registered faces. Only detecting unknown users.")

    cap = cv2.VideoCapture(0)
    observed_faces = {}  # Track consistent recognition

    while True:
        ret, frame = cap.read()
        if not ret:
            continue

        recognized_faces = recognize_faces(frame)
        for name, (x1, y1, x2, y2) in recognized_faces:
            if not known_face_encodings:
                name = "Unknown"  # Force unknown if no registered faces

            if name != "Unknown":
                if name in observed_faces:
                    observed_faces[name]["count"] += 1
                    observed_faces[name]["last_seen"] = time.time()
                else:
                    observed_faces[name] = {"count": 1, "last_seen": time.time()}

                # Ensure face is recognized for at least 10 seconds before marking attendance
                if observed_faces[name]["count"] >= 10:  # 10s continuous check
                    mark_attendance(name)
                    print(f"✅ {name} confirmed and attendance marked!")

            # Set box color (Red for Unknown, Green for Recognized)
            box_color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)

            # Draw rectangle around the face
            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, box_color, 2)

        cv2.imshow("Face Attendance System", frame)

        # Remove stale records after 10 seconds
        current_time = time.time()
        observed_faces = {k: v for k, v in observed_faces.items() if current_time - v["last_seen"] < 15} 

        # Check for ESC key press (ASCII 27)
        key = cv2.waitKey(1) & 0xFF
        if key == 27:  # 27 is the ASCII code for ESC
            print("🚪 ESC key pressed. Exiting webcam & saving attendance...")
            close_session()  # Close the current session when ESC is pressed
            break

    cap.release()
    cv2.destroyAllWindows()

    # Force page to refresh after attendance is marked
    return redirect(url_for("home"))

if __name__ == "__main__":
    app.run(debug=True, port=5000)
