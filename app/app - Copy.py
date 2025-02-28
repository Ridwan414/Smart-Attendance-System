import cv2
import os
from flask import Flask, request, render_template, redirect, url_for, jsonify
from datetime import date, datetime
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import GridSearchCV
import pandas as pd
import joblib
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

nimgs = 100
imgBackground = cv2.imread("background.png")

datetoday = date.today().strftime("%m_%d_%y")
datetoday2 = date.today().strftime("%d-%B-%Y")

face_detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

if not os.path.isdir('Attendance'):
    os.makedirs('Attendance')
if not os.path.isdir('static'):
    os.makedirs('static')
if not os.path.isdir('static/faces'):
    os.makedirs('static/faces')

# Excel sheet related code
current_wb = None
current_sheet = None
current_file = None


def get_today_filename():
    today = datetime.now().strftime('%d_%m_%Y')
    return f"Attendance-{today}.xlsx"


def get_period_name(index):
    if index % 10 == 1 and index % 100 != 11:
        suffix = "st"
    elif index % 10 == 2 and index % 100 != 12:
        suffix = "nd"
    elif index % 10 == 3 and index % 100 != 13:
        suffix = "rd"
    else:
        suffix = "th"
    return f"{index}{suffix} Period"


def start_process():
    global current_wb, current_sheet, current_file
    try:
        current_file = get_today_filename()
        if os.path.exists(f'Attendance/{current_file}'):
            current_wb = load_workbook(f'Attendance/{current_file}')
            sheet_count = len(current_wb.sheetnames)
        else:
            current_wb = Workbook()
            default_sheet = current_wb.active
            current_wb.remove(default_sheet)
            sheet_count = 0

        period_name = get_period_name(sheet_count + 1)

        current_sheet = current_wb.create_sheet(period_name)
        current_sheet['A1'] = "Name"
        current_sheet['B1'] = "ID"
        current_sheet['C1'] = "Time"

        current_wb.save(f'Attendance/{current_file}')
    except Exception as e:
        print(f"Error starting process: {e}")


def totalreg():
    return len(os.listdir('static/faces'))


def extract_faces(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    face_points = face_detector.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(50, 50))
    return face_points


def identify_face(facearray):
    if not os.path.exists('static/face_recognition_model.pkl'):
        return ["Unknown"]

    model = joblib.load('static/face_recognition_model.pkl')
    try:
        distances, indices = model.kneighbors(facearray, n_neighbors=1)
        threshold = 5000  # Adjust as needed
        if distances[0][0] > threshold:
            return ["Unknown"]
        return model.predict(facearray)
    except:
        return ["Unknown"]


def train_model():
    faces = []
    labels = []
    userlist = os.listdir('static/faces')
    for user in userlist:
        for imgname in os.listdir(f'static/faces/{user}'):
            img = cv2.imread(f'static/faces/{user}/{imgname}')
            resized_face = cv2.resize(img, (50, 50))
            faces.append(resized_face.ravel())
            labels.append(user)
    faces = np.array(faces)

    # Hyperparameter tuning using GridSearchCV
    param_grid = {
        'n_neighbors': [3, 5, 7, 9, 11],
        'weights': ['uniform', 'distance'],
        'algorithm': ['auto', 'ball_tree', 'kd_tree', 'brute'],
        'metric': ['euclidean', 'manhattan', 'chebyshev']
    }

    knn = KNeighborsClassifier()
    grid_search = GridSearchCV(knn, param_grid, cv=3, n_jobs=-1, verbose=2, scoring='accuracy')
    grid_search.fit(faces, labels)

    # Best parameters found by GridSearchCV
    best_knn = grid_search.best_estimator_

    # Save the best model
    joblib.dump(best_knn, 'static/face_recognition_model.pkl')


def extract_attendance():
    current_file = get_today_filename()
    if os.path.exists(f'Attendance/{current_file}'):
        current_wb = load_workbook(f'Attendance/{current_file}')
        current_sheet = current_wb.sheetnames[-1]
        sheet = current_wb[current_sheet]

        names = []
        rolls = []
        times = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            names.append(row[0])
            rolls.append(row[1])
            times.append(row[2])
        l = len(names)
        return names, rolls, times, l
    else:
        return [], [], [], 0


def add_attendance(name):
    username = name.split('_')[0]
    userid = name.split('_')[1]
    current_time = datetime.now().strftime("%H:%M:%S")

    row = [username, userid, current_time]
    current_sheet.append(row)
    current_wb.save(f'Attendance/{current_file}')


def getallusers():
    userlist = os.listdir('static/faces')
    names = []
    rolls = []
    l = len(userlist)
    for i in userlist:
        name, roll = i.split('_')
        names.append(name)
        rolls.append(roll)
    return userlist, names, rolls, l


@app.route('/')
def home():
    names, rolls, times, l = extract_attendance()
    return render_template('home.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(),
                           datetoday2=datetoday2)


@app.route('/take_attendance', methods=['GET'])
def take_attendance():
    return redirect(url_for('start'))


@app.route('/start', methods=['GET'])
def start():
    names, rolls, times, l = extract_attendance()

    if 'face_recognition_model.pkl' not in os.listdir('static'):
        return render_template('home.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(),
                               datetoday2=datetoday2, mess='No trained model found. Please add a face to continue.')

    start_process()

    cap = cv2.VideoCapture(0)
    added_faces = set()
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        faces = extract_faces(frame)
        for (x, y, w, h) in faces:
            face = cv2.resize(frame[y:y + h, x:x + w], (50, 50))
            identified_person = identify_face(face.reshape(1, -1))[0]

            if identified_person != "Unknown" and identified_person not in added_faces:
                add_attendance(identified_person)
                added_faces.add(identified_person)

            if identified_person == "Unknown":
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.putText(frame, identified_person, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
            else:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, identified_person, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

        imgBackground[162:162 + 480, 55:55 + 640] = frame
        cv2.imshow('Attendance', imgBackground)
        if cv2.waitKey(1) == 27:
            break

    cap.release()
    cv2.destroyAllWindows()
    names, rolls, times, l = extract_attendance()
    return render_template('home.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(),
                           datetoday2=datetoday2)


@app.route('/end_process', methods=['POST'])
def end_process():
    global current_wb, current_sheet
    try:
        current_wb.save(f'Attendance/{current_file}')
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route('/add', methods=['GET', 'POST'])
def add():
    newusername = request.form['newusername']
    newuserid = request.form['newuserid']
    userimagefolder = f'static/faces/{newusername}_{newuserid}'
    if not os.path.isdir(userimagefolder):
        os.makedirs(userimagefolder)

    i, j = 0, 0
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        faces = extract_faces(frame)
        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 20), 2)
            cv2.putText(frame, f'Images Captured: {i}/{nimgs}', (30, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 20), 2, cv2.LINE_AA)
            if j % 5 == 0:
                name = f'{newusername}_{i}.jpg'
                cv2.imwrite(f'{userimagefolder}/{name}', frame[y:y + h, x:x + w])
                i += 1
            j += 1
        if j == nimgs * 5:
            break
        cv2.imshow('Adding New User', frame)
        if cv2.waitKey(1) == 27:
            break

    cap.release()
    cv2.destroyAllWindows()
    train_model()
    names, rolls, times, l = extract_attendance()
    return render_template('home.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(),
                           datetoday2=datetoday2)


if __name__ == '__main__':
    app.run(debug=True)
